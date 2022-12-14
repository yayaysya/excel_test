
# 导入的库资源
import openpyxl
import logging
import os

from priv_class import CTRL
from priv_class import init_log

# 初始化日志模块
cur_path = os.getcwd()
path = os.path.join(cur_path, 'log.txt')
log = init_log(path)

# 初始化全局变量
ctrl=CTRL()

# 初始化一些宏
SYS_OK = 0
SYS_ERR = -1

'''
func: 保存表格
'''
def excel_save():
    new_dev_cfg_excel_path = os.path.join(cur_path, 'dev_cfg_h8_new.xlsx')
    ctrl.sys_dev_cfg_excel.save(new_dev_cfg_excel_path)

'''
func:数据校验: 
1. 校验ipp的表里面, 是否存在相同的两个dev_type, 但是bss_type却不同
'''
def execl_ipp_dev_type_check():
    dev_row = ctrl.ipp_dev_type_cell.row + 1
    dev_cow = ctrl.ipp_dev_type_cell.column
    dev_cow_letter = ctrl.ipp_dev_type_cell.column_letter
    ipp_max_row = ctrl.ipp_bss_cfg_sheet.max_row
    while dev_row <= ipp_max_row:
        next_dev_row = dev_row+1
        while next_dev_row <= ipp_max_row:
            if ctrl.ipp_bss_cfg_sheet.cell(dev_row, dev_cow).value \
                    == ctrl.ipp_bss_cfg_sheet.cell(next_dev_row, dev_cow).value:
                for bss_cell in ctrl.ipp_bss_cell:
                    if ctrl.ipp_bss_cfg_sheet.cell(dev_row, bss_cell.column).value \
                        != ctrl.ipp_bss_cfg_sheet.cell(next_dev_row, bss_cell.column).value:
                        log.error(f"ERROR: file:{ctrl.ipp_bss_cfg_filename} "
                                  f"dev_type[{dev_row}{dev_cow_letter}],[{next_dev_row}{dev_cow_letter}]same, "
                                  f"but bss_type is not same")
                        return SYS_ERR

            next_dev_row += 1

        dev_row += 1
    return SYS_OK



'''
func: 数据匹配
'''

def excel_match():

    sys_max_row = ctrl.sys_dev_cfg_sheet.max_row
    ipp_max_row = ctrl.ipp_bss_cfg_sheet.max_row

    ipp_row = ctrl.ipp_bss_cell[0].row + 1
    while ipp_row <= ipp_max_row:
        ipp_dev_type_cell = ctrl.ipp_bss_cfg_sheet.cell(ipp_row, ctrl.ipp_dev_type_cell.column)
        sys_row = ctrl.sys_dev_type_cell.row + 1
        while sys_row <= sys_max_row:
            sys_dev_type_cell = ctrl.sys_dev_cfg_sheet.cell(sys_row, ctrl.sys_dev_type_cell.column)

            if (sys_dev_type_cell.value == ipp_dev_type_cell.value):
                for sys_cell in ctrl.sys_bss0_cell:
                    for ipp_cell in ctrl.ipp_bss_cell:
                        if (sys_cell.value == ipp_cell.value):
                            ctrl.sys_dev_cfg_sheet.cell(sys_row, sys_cell.column).value \
                                = ctrl.ipp_bss_cfg_sheet.cell(ipp_row, ipp_cell.column).value
                            #print(ctrl.sys_dev_cfg_sheet.cell(sys_row, sys_cell.column).value)
                            break #匹配到bss名称
                break#匹配到dev_type

            sys_row += 1
        if (sys_row > sys_max_row):
            log.error(f"ERROR! Please check: sys_file: {ctrl.sys_dev_cfg_filename} can't find ipp_file dev_type:{ipp_dev_type_cell.value}")
            return SYS_ERR
        ipp_row += 1

'''
func: 找到对应的cell
'''
def excel_find_cell():
    ctrl.sys_dev_cfg_excel = openpyxl.load_workbook(ctrl.sys_dev_cfg_filename)
    ctrl.sys_dev_cfg_sheet = ctrl.sys_dev_cfg_excel.active  #获取当前活跃的sheet

    #找到sys_dev_type_cell  sys_bss0_cell
    sys_dev_type_cell = SYS_ERR
    sys_bss0_cell = []
    for row in ctrl.sys_dev_cfg_sheet:
        for cell in row:
            if cell.value == 'dev_type':
                sys_dev_type_cell = cell
            if (cell.value == 'bss0' or cell.value == 'bss1'):
                sys_bss0_cell.append(cell)
    #log.info(f"sys_dev_type_cell:{sys_dev_type_cell}, sys_bss0_cell:{sys_bss0_cell}")

    if (SYS_ERR == sys_dev_type_cell or sys_bss0_cell == []):
        log.error(f"file:{ctrl.sys_dev_cfg_filename} can't find key word(dev_type, bss0, bss1), please check\n")
        return SYS_ERR

    ctrl.sys_dev_type_cell = sys_dev_type_cell
    ctrl.sys_bss0_cell = sys_bss0_cell


    #找到ipp_bss_cell
    ipp_dev_type_cell = SYS_ERR
    ipp_bss_cell = []
    ctrl.ipp_bss_cfg_excel = openpyxl.load_workbook(ctrl.ipp_bss_cfg_filename)
    ctrl.ipp_bss_cfg_sheet = ctrl.ipp_bss_cfg_excel.active  #获取当前活跃的sheet
    for row in ctrl.ipp_bss_cfg_sheet:
        for cell in row:
            if cell.value == 'dev_type':
                ipp_dev_type_cell = cell
            if (cell.value == 'bss0' or cell.value == 'bss1'):
                ipp_bss_cell.append(cell)

    if (SYS_ERR == ipp_dev_type_cell or ipp_bss_cell == []):
        log.error(f"file:{ctrl.ipp_bss_cfg_filename} can't find key word(dev_type, bss0, bss1), please check\n")
        return SYS_ERR

    ctrl.ipp_dev_type_cell = ipp_dev_type_cell
    ctrl.ipp_bss_cell = ipp_bss_cell
    #log.info(f"ipp_bss_cell:{ctrl.ipp_bss_cell}")



'''
func: 打开配置文件
'''
def excel_get_filename():
    # 打开配置文件
    f = open('config.txt', 'r', encoding='utf8')
    lines = f.readlines()
    ctrl.sys_dev_cfg_filename = lines[0].split(':')[1].split('\n')[0]
    ctrl.sys_dev_cfg_filename = os.path.join(cur_path, ctrl.sys_dev_cfg_filename)
    ctrl.ipp_bss_cfg_filename = lines[1].split(':')[1].split('\n')[0]
    ctrl.ipp_bss_cfg_filename = os.path.join(cur_path, ctrl.ipp_bss_cfg_filename)
    log.info(f"sys_dev_cfg_filename:{ctrl.sys_dev_cfg_filename}")
    log.info(f"ipp_bss_cfg_filename:{ctrl.ipp_bss_cfg_filename}")
    f.close()

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    log.info('start to proc excel data')
    #打开配置文件
    ret = excel_get_filename()
    if (SYS_ERR == ret):exit(-1)

    #获取关键cell
    ret = excel_find_cell()
    if (SYS_ERR == ret): exit(-1)

    #结果校验
    ret = execl_ipp_dev_type_check()
    if (SYS_ERR == ret): exit(-1)

    #表格处理
    ret = excel_match()
    if (SYS_ERR == ret): exit(-1)

    #表格保存
    ret = excel_save()
    if (SYS_ERR == ret): exit(-1)

    log.info('proc excel data success!')


